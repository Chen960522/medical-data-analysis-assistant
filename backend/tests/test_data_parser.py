"""Tests for data parsing and quality detection service.

Tests cover: CSV/Excel/JSON parsing, column type detection,
missing value detection, format error detection, and quality summary generation.
Requirements: 2.1-2.6
"""

import json
import io

import pytest

from app.services.data_parser import (
    ParseResult,
    MissingValueReport,
    FormatError,
    QualitySummary,
    ColumnQualityInfo,
    parse_file,
    detect_column_types,
    detect_missing_values,
    detect_format_errors,
    generate_quality_summary,
)


# --- CSV Parsing Tests ---


class TestParseCSV:
    """Tests for CSV file parsing."""

    def test_parse_csv_utf8(self):
        """Test parsing a standard UTF-8 CSV file."""
        content = b"name,age,score\nAlice,30,90\nBob,25,85\n"
        result = parse_file(content, "csv")

        assert isinstance(result, ParseResult)
        assert result.columns == ["name", "age", "score"]
        assert result.total_rows == 2
        assert result.total_columns == 3
        assert result.rows[0]["name"] == "Alice"
        assert result.rows[1]["age"] == "25"

    def test_parse_csv_utf8_bom(self):
        """Test parsing a CSV file with UTF-8 BOM."""
        bom = b"\xef\xbb\xbf"
        content = bom + b"name,age\nAlice,30\nBob,25\n"
        result = parse_file(content, "csv")

        assert result.columns == ["name", "age"]
        assert result.total_rows == 2
        assert result.rows[0]["name"] == "Alice"

    def test_parse_csv_empty_file(self):
        """Test parsing an empty CSV (header only)."""
        content = b"name,age,score\n"
        result = parse_file(content, "csv")

        assert result.columns == ["name", "age", "score"]
        assert result.total_rows == 0
        assert result.total_columns == 3

    def test_parse_csv_single_row(self):
        """Test parsing a CSV with a single data row."""
        content = b"id,value\n1,hello\n"
        result = parse_file(content, "csv")

        assert result.total_rows == 1
        assert result.rows[0]["id"] == "1"
        assert result.rows[0]["value"] == "hello"

    def test_parse_csv_single_column(self):
        """Test parsing a CSV with a single column."""
        content = b"name\nAlice\nBob\nCharlie\n"
        result = parse_file(content, "csv")

        assert result.columns == ["name"]
        assert result.total_rows == 3
        assert result.total_columns == 1

    def test_parse_csv_with_missing_values(self):
        """Test parsing a CSV with missing values."""
        content = b"name,age,score\nAlice,30,90\nBob,,85\nCharlie,25,\n"
        result = parse_file(content, "csv")

        assert result.total_rows == 3
        assert result.rows[1]["age"] == ""
        assert result.rows[2]["score"] == ""


# --- Excel Parsing Tests ---


class TestParseExcel:
    """Tests for Excel file parsing."""

    def test_parse_xlsx(self):
        """Test parsing a basic Excel file."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "age", "score"])
        ws.append(["Alice", 30, 90])
        ws.append(["Bob", 25, 85])

        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()

        result = parse_file(content, "xlsx")

        assert result.columns == ["name", "age", "score"]
        assert result.total_rows == 2
        assert result.total_columns == 3
        assert result.rows[0]["name"] == "Alice"
        assert result.rows[0]["age"] == 30

    def test_parse_xlsx_empty(self):
        """Test parsing an empty Excel file (no data rows)."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["col1", "col2"])

        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()

        result = parse_file(content, "xlsx")

        assert result.columns == ["col1", "col2"]
        assert result.total_rows == 0

    def test_parse_xlsx_with_none_header(self):
        """Test parsing Excel with None values in header."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", None, "score"])
        ws.append(["Alice", 30, 90])

        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()

        result = parse_file(content, "xlsx")

        assert "column_1" in result.columns
        assert result.total_rows == 1


# --- JSON Parsing Tests ---


class TestParseJSON:
    """Tests for JSON file parsing."""

    def test_parse_json_array_of_objects(self):
        """Test parsing JSON as array of objects."""
        data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob", "age": 25},
        ]
        content = json.dumps(data).encode("utf-8")
        result = parse_file(content, "json")

        assert result.columns == ["name", "age"]
        assert result.total_rows == 2
        assert result.rows[0]["name"] == "Alice"

    def test_parse_json_with_data_key(self):
        """Test parsing JSON with nested 'data' key."""
        data = {
            "data": [
                {"id": 1, "value": "a"},
                {"id": 2, "value": "b"},
            ]
        }
        content = json.dumps(data).encode("utf-8")
        result = parse_file(content, "json")

        assert result.columns == ["id", "value"]
        assert result.total_rows == 2

    def test_parse_json_single_object(self):
        """Test parsing JSON as a single object."""
        data = {"name": "Alice", "age": 30, "score": 90}
        content = json.dumps(data).encode("utf-8")
        result = parse_file(content, "json")

        assert result.total_rows == 1
        assert result.total_columns == 3

    def test_parse_json_empty_array(self):
        """Test parsing an empty JSON array."""
        content = b"[]"
        result = parse_file(content, "json")

        assert result.columns == []
        assert result.total_rows == 0

    def test_parse_json_utf8_bom(self):
        """Test parsing JSON with UTF-8 BOM."""
        bom = b"\xef\xbb\xbf"
        data = [{"name": "Alice"}]
        content = bom + json.dumps(data).encode("utf-8")
        result = parse_file(content, "json")

        assert result.total_rows == 1
        assert result.rows[0]["name"] == "Alice"


# --- Column Type Detection Tests ---


class TestDetectColumnTypes:
    """Tests for column type auto-detection."""

    def test_detect_numeric_integers(self):
        """Test detection of integer numeric columns."""
        columns = ["value"]
        rows = [{"value": "1"}, {"value": "2"}, {"value": "3"}, {"value": "4"}, {"value": "5"}]
        types = detect_column_types(columns, rows)
        assert types["value"] == "numeric"

    def test_detect_numeric_floats(self):
        """Test detection of float numeric columns."""
        columns = ["value"]
        rows = [{"value": "1.5"}, {"value": "2.7"}, {"value": "3.14"}, {"value": "0.5"}, {"value": "9.9"}]
        types = detect_column_types(columns, rows)
        assert types["value"] == "numeric"

    def test_detect_numeric_native_types(self):
        """Test detection of native numeric types (int/float)."""
        columns = ["value"]
        rows = [{"value": 1}, {"value": 2.5}, {"value": 3}, {"value": 4.0}, {"value": 5}]
        types = detect_column_types(columns, rows)
        assert types["value"] == "numeric"

    def test_detect_categorical(self):
        """Test detection of categorical columns (low unique ratio)."""
        columns = ["status"]
        # 10 rows with only 2 unique values -> ratio = 2/10 = 0.2 < 0.3
        rows = [{"status": "active"} for _ in range(5)] + [{"status": "inactive"} for _ in range(5)]
        types = detect_column_types(columns, rows)
        assert types["status"] == "categorical"

    def test_detect_date_iso_format(self):
        """Test detection of date columns with ISO format."""
        columns = ["date"]
        rows = [
            {"date": "2024-01-01"},
            {"date": "2024-02-15"},
            {"date": "2024-03-20"},
            {"date": "2024-04-10"},
            {"date": "2024-05-05"},
        ]
        types = detect_column_types(columns, rows)
        assert types["date"] == "date"

    def test_detect_date_slash_format(self):
        """Test detection of date columns with slash format."""
        columns = ["date"]
        rows = [
            {"date": "2024/01/01"},
            {"date": "2024/02/15"},
            {"date": "2024/03/20"},
            {"date": "2024/04/10"},
            {"date": "2024/05/05"},
        ]
        types = detect_column_types(columns, rows)
        assert types["date"] == "date"

    def test_detect_text(self):
        """Test detection of text columns (high unique ratio, not numeric/date)."""
        columns = ["description"]
        rows = [
            {"description": "Patient shows improvement"},
            {"description": "No significant changes"},
            {"description": "Mild symptoms observed"},
            {"description": "Treatment ongoing"},
            {"description": "Follow-up required"},
        ]
        types = detect_column_types(columns, rows)
        assert types["description"] == "text"

    def test_detect_empty_column(self):
        """Test detection of column with all missing values."""
        columns = ["empty"]
        rows = [{"empty": None}, {"empty": ""}, {"empty": None}]
        types = detect_column_types(columns, rows)
        assert types["empty"] == "text"

    def test_detect_multiple_columns(self):
        """Test detection of multiple columns at once."""
        columns = ["id", "name", "age", "date"]
        rows = [
            {"id": "1", "name": "Alice", "age": "30", "date": "2024-01-01"},
            {"id": "2", "name": "Bob", "age": "25", "date": "2024-02-01"},
            {"id": "3", "name": "Charlie", "age": "35", "date": "2024-03-01"},
            {"id": "4", "name": "Diana", "age": "28", "date": "2024-04-01"},
            {"id": "5", "name": "Eve", "age": "32", "date": "2024-05-01"},
        ]
        types = detect_column_types(columns, rows)
        assert types["age"] == "numeric"
        assert types["date"] == "date"


# --- Missing Value Detection Tests ---


class TestDetectMissingValues:
    """Tests for missing value detection and reporting."""

    def test_no_missing_values(self):
        """Test dataset with no missing values."""
        columns = ["a", "b"]
        rows = [{"a": 1, "b": 2}, {"a": 3, "b": 4}]
        report = detect_missing_values(columns, rows)

        assert isinstance(report, MissingValueReport)
        assert report.total_missing == 0
        assert report.positions == []
        assert report.overall_percentage == 0.0

    def test_none_values(self):
        """Test detection of None as missing."""
        columns = ["a", "b"]
        rows = [{"a": None, "b": 2}, {"a": 3, "b": None}]
        report = detect_missing_values(columns, rows)

        assert report.total_missing == 2
        assert (0, "a") in report.positions
        assert (1, "b") in report.positions

    def test_empty_string_values(self):
        """Test detection of empty strings as missing."""
        columns = ["name", "age"]
        rows = [{"name": "Alice", "age": ""}, {"name": "", "age": "30"}]
        report = detect_missing_values(columns, rows)

        assert report.total_missing == 2
        assert (0, "age") in report.positions
        assert (1, "name") in report.positions

    def test_nan_values(self):
        """Test detection of NaN as missing."""
        columns = ["value"]
        rows = [{"value": float("nan")}, {"value": 1.0}]
        report = detect_missing_values(columns, rows)

        assert report.total_missing == 1
        assert (0, "value") in report.positions

    def test_per_column_counts(self):
        """Test per-column missing value counts."""
        columns = ["a", "b", "c"]
        rows = [
            {"a": None, "b": 1, "c": None},
            {"a": None, "b": None, "c": 3},
            {"a": 1, "b": 2, "c": 3},
        ]
        report = detect_missing_values(columns, rows)

        assert report.per_column["a"] == 2
        assert report.per_column["b"] == 1
        assert report.per_column["c"] == 1
        assert report.total_missing == 4
        assert report.total_cells == 9

    def test_overall_percentage(self):
        """Test overall missing percentage calculation."""
        columns = ["a", "b"]
        rows = [{"a": None, "b": None}, {"a": 1, "b": 2}]
        report = detect_missing_values(columns, rows)

        # 2 missing out of 4 cells = 50%
        assert report.overall_percentage == 50.0

    def test_empty_dataset(self):
        """Test missing value detection on empty dataset."""
        columns = ["a", "b"]
        rows = []
        report = detect_missing_values(columns, rows)

        assert report.total_missing == 0
        assert report.overall_percentage == 0.0


# --- Format Error Detection Tests ---


class TestDetectFormatErrors:
    """Tests for format error detection."""

    def test_csv_no_errors(self):
        """Test CSV with no format errors."""
        content = b"a,b,c\n1,2,3\n4,5,6\n"
        errors = detect_format_errors(content, "csv")
        assert errors == []

    def test_csv_inconsistent_columns(self):
        """Test CSV with inconsistent column counts."""
        content = b"a,b,c\n1,2,3\n4,5\n6,7,8,9\n"
        errors = detect_format_errors(content, "csv")

        assert len(errors) >= 2
        assert any("Inconsistent column count" in e.description for e in errors)

    def test_csv_encoding_error(self):
        """Test CSV with encoding issues."""
        # Invalid UTF-8 bytes
        content = b"name,value\n\xff\xfe,test\n"
        errors = detect_format_errors(content, "csv")

        assert any("encoding" in e.description.lower() or "UTF-8" in e.description for e in errors)

    def test_json_parse_error(self):
        """Test JSON with parse errors."""
        content = b'{"name": "Alice", invalid}'
        errors = detect_format_errors(content, "json")

        assert len(errors) >= 1
        assert any("parse error" in e.description.lower() or "JSON" in e.description for e in errors)

    def test_json_inconsistent_schema(self):
        """Test JSON with inconsistent schemas between records."""
        data = [
            {"name": "Alice", "age": 30},
            {"name": "Bob"},  # missing "age"
        ]
        content = json.dumps(data).encode("utf-8")
        errors = detect_format_errors(content, "json")

        assert len(errors) >= 1
        assert any("Missing keys" in e.description for e in errors)

    def test_json_non_dict_entries(self):
        """Test JSON with non-object entries in array."""
        data = [{"name": "Alice"}, "not an object", {"name": "Bob"}]
        content = json.dumps(data).encode("utf-8")
        errors = detect_format_errors(content, "json")

        assert len(errors) >= 1
        assert any("not an object" in e.description for e in errors)

    def test_excel_format_errors(self):
        """Test Excel format error detection (merged cells)."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["name", "age", "score"])
        ws.append(["Alice", 30, 90])
        ws.append(["Bob", 25, 85])
        ws.merge_cells("A4:B4")

        buffer = io.BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()

        errors = detect_format_errors(content, "xlsx")

        assert len(errors) >= 1
        assert any("Merged cells" in e.description for e in errors)

    def test_unsupported_format(self):
        """Test unsupported format returns error."""
        errors = detect_format_errors(b"content", "xml")
        assert len(errors) == 1
        assert "Unsupported" in errors[0].description


# --- Quality Summary Tests ---


class TestGenerateQualitySummary:
    """Tests for quality summary generation."""

    def test_basic_quality_summary(self):
        """Test quality summary with clean data."""
        columns = ["name", "age", "score"]
        rows = [
            {"name": "Alice", "age": 30, "score": 90},
            {"name": "Bob", "age": 25, "score": 85},
            {"name": "Charlie", "age": 35, "score": 92},
        ]
        summary = generate_quality_summary(columns, rows)

        assert isinstance(summary, QualitySummary)
        assert summary.total_rows == 3
        assert summary.total_columns == 3
        assert summary.missing_value_percentage == 0.0
        assert summary.completeness_score == 100.0
        assert len(summary.columns) == 3

    def test_quality_summary_with_missing(self):
        """Test quality summary with missing values."""
        columns = ["a", "b"]
        rows = [
            {"a": 1, "b": None},
            {"a": None, "b": 2},
            {"a": 3, "b": 3},
        ]
        summary = generate_quality_summary(columns, rows)

        assert summary.total_rows == 3
        assert summary.missing_value_percentage > 0
        assert summary.completeness_score < 100.0
        # 2 missing out of 6 cells = 33.33%
        assert abs(summary.missing_value_percentage - 33.33) < 0.1

    def test_quality_summary_column_info(self):
        """Test per-column quality info in summary."""
        columns = ["id", "status"]
        rows = [
            {"id": "1", "status": "active"},
            {"id": "2", "status": "active"},
            {"id": "3", "status": "inactive"},
            {"id": "4", "status": "active"},
            {"id": "5", "status": None},
        ]
        summary = generate_quality_summary(columns, rows)

        # Find status column info
        status_col = next(c for c in summary.columns if c.name == "status")
        assert status_col.missing_count == 1
        assert status_col.missing_percentage == 20.0
        assert status_col.unique_count == 2  # "active" and "inactive"
        assert len(status_col.sample_values) <= 5

    def test_quality_summary_detects_types(self):
        """Test that quality summary includes detected types."""
        columns = ["num", "date_col", "text_col"]
        rows = [
            {"num": "1.5", "date_col": "2024-01-01", "text_col": "hello world"},
            {"num": "2.7", "date_col": "2024-02-01", "text_col": "foo bar"},
            {"num": "3.0", "date_col": "2024-03-01", "text_col": "baz qux"},
            {"num": "4.2", "date_col": "2024-04-01", "text_col": "test data"},
            {"num": "5.1", "date_col": "2024-05-01", "text_col": "sample text"},
        ]
        summary = generate_quality_summary(columns, rows)

        type_map = {c.name: c.dtype for c in summary.columns}
        assert type_map["num"] == "numeric"
        assert type_map["date_col"] == "date"

    def test_quality_summary_empty_dataset(self):
        """Test quality summary with empty dataset."""
        columns = ["a", "b"]
        rows = []
        summary = generate_quality_summary(columns, rows)

        assert summary.total_rows == 0
        assert summary.total_columns == 2
        assert summary.missing_value_percentage == 0.0
        assert summary.completeness_score == 100.0

    def test_quality_summary_sample_values(self):
        """Test that sample values are limited to 5 unique values."""
        columns = ["val"]
        rows = [{"val": f"item_{i}"} for i in range(20)]
        summary = generate_quality_summary(columns, rows)

        val_col = summary.columns[0]
        assert len(val_col.sample_values) == 5


# --- Edge Cases ---


class TestEdgeCases:
    """Tests for edge cases."""

    def test_unsupported_format_raises(self):
        """Test that unsupported format raises ValueError."""
        with pytest.raises(ValueError, match="Unsupported"):
            parse_file(b"content", "xml")

    def test_parse_completely_empty_csv(self):
        """Test parsing a completely empty CSV (no header)."""
        content = b""
        result = parse_file(content, "csv")
        assert result.columns == []
        assert result.total_rows == 0

    def test_missing_key_in_row(self):
        """Test missing value detection when key is absent from row dict."""
        columns = ["a", "b"]
        rows = [{"a": 1}]  # "b" key is missing
        report = detect_missing_values(columns, rows)

        assert report.total_missing == 1
        assert (0, "b") in report.positions
