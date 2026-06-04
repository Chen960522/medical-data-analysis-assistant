"""Data parsing and quality detection service.

Provides file parsing (CSV/Excel/JSON), column type auto-detection,
missing value detection, format error detection, and quality summary generation.

Requirements: 2.1-2.6
"""

import csv
import io
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any


@dataclass
class ParseResult:
    """Result of parsing a file into structured tabular data."""

    columns: list[str]
    rows: list[dict[str, Any]]
    total_rows: int
    total_columns: int


@dataclass
class MissingValueReport:
    """Report of missing values in the dataset."""

    positions: list[tuple[int, str]]  # (row_index, column_name)
    per_column: dict[str, int]  # column_name -> missing count
    total_missing: int
    total_cells: int
    overall_percentage: float


@dataclass
class FormatError:
    """A single format error detected in the file."""

    row: int  # 1-based row number
    column: str | None  # column name if applicable
    description: str


@dataclass
class ColumnQualityInfo:
    """Quality information for a single column."""

    name: str
    dtype: str
    missing_count: int
    missing_percentage: float
    unique_count: int
    sample_values: list[Any]


@dataclass
class QualitySummary:
    """Comprehensive data quality summary."""

    total_rows: int
    total_columns: int
    missing_value_percentage: float
    completeness_score: float  # percentage of non-missing cells
    columns: list[ColumnQualityInfo]


# --- File Parsing ---


def parse_file(content: bytes, file_format: str) -> ParseResult:
    """Parse file content into structured tabular data.

    Args:
        content: Raw file bytes.
        file_format: File format string (e.g., "csv", "xlsx", "xls", "json").

    Returns:
        ParseResult with columns, rows, and counts.

    Raises:
        ValueError: If the file format is unsupported or content cannot be parsed.
    """
    fmt = file_format.lstrip(".").lower()

    if fmt == "csv":
        columns, rows, total = _parse_csv(content)
    elif fmt in ("xlsx", "xls"):
        columns, rows, total = _parse_excel(content)
    elif fmt == "json":
        columns, rows, total = _parse_json(content)
    else:
        raise ValueError(f"Unsupported file format: {file_format}")

    return ParseResult(
        columns=columns,
        rows=rows,
        total_rows=total,
        total_columns=len(columns),
    )


def _parse_csv(content: bytes) -> tuple[list[str], list[dict[str, Any]], int]:
    """Parse CSV content. Tries UTF-8 first, then UTF-8 with BOM."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    reader = csv.DictReader(io.StringIO(text))
    columns = list(reader.fieldnames or [])
    all_rows = list(reader)
    return columns, all_rows, len(all_rows)


def _parse_excel(content: bytes) -> tuple[list[str], list[dict[str, Any]], int]:
    """Parse Excel content using openpyxl."""
    try:
        import openpyxl
    except ImportError:
        raise ValueError("Excel parsing library (openpyxl) not available")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return [], [], 0

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        wb.close()
        return [], [], 0

    columns = [str(c) if c is not None else f"column_{i}" for i, c in enumerate(header_row)]
    all_rows = []
    for row in rows_iter:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(columns):
                row_dict[columns[i]] = val
        all_rows.append(row_dict)

    wb.close()
    return columns, all_rows, len(all_rows)


def _parse_json(content: bytes) -> tuple[list[str], list[dict[str, Any]], int]:
    """Parse JSON content (array of objects or dict with 'data' key)."""
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")

    data = json.loads(text)

    if isinstance(data, list):
        if len(data) == 0:
            return [], [], 0
        columns = list(data[0].keys()) if isinstance(data[0], dict) else []
        all_rows = data if all(isinstance(r, dict) for r in data) else []
        return columns, all_rows, len(all_rows)
    elif isinstance(data, dict):
        if "data" in data and isinstance(data["data"], list):
            records = data["data"]
            columns = list(records[0].keys()) if records and isinstance(records[0], dict) else []
            return columns, records, len(records)
        # Single record
        columns = list(data.keys())
        return columns, [data], 1
    else:
        return [], [], 0


# --- Column Type Detection ---


def detect_column_types(columns: list[str], rows: list[dict[str, Any]]) -> dict[str, str]:
    """Detect data types for each column.

    Types: "numeric", "categorical", "date", "text".

    Args:
        columns: List of column names.
        rows: List of row dictionaries.

    Returns:
        Mapping of column name to detected type string.
    """
    result = {}
    for col in columns:
        values = [row.get(col) for row in rows]
        result[col] = _detect_single_column_type(values)
    return result


def _detect_single_column_type(values: list[Any]) -> str:
    """Detect the type of a single column from its values."""
    non_null = [v for v in values if v is not None and str(v).strip() != ""]
    if not non_null:
        return "text"

    sample = non_null[:100]
    sample_size = len(sample)

    # Check numeric
    numeric_count = 0
    for v in sample:
        if isinstance(v, (int, float)):
            numeric_count += 1
        elif isinstance(v, str):
            try:
                float(v)
                numeric_count += 1
            except (ValueError, TypeError):
                pass

    if numeric_count > sample_size * 0.8:
        return "numeric"

    # Check date
    date_count = 0
    for v in sample:
        if isinstance(v, (datetime, date)):
            date_count += 1
        elif isinstance(v, str):
            if _is_date_string(v):
                date_count += 1

    if date_count > sample_size * 0.8:
        return "date"

    # Check categorical (few unique values relative to total)
    unique_ratio = len(set(str(v) for v in sample)) / sample_size
    if unique_ratio < 0.3:
        return "categorical"

    return "text"


_DATE_PATTERNS = [
    re.compile(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}"),  # YYYY-MM-DD or YYYY/MM/DD
    re.compile(r"^\d{1,2}[-/]\d{1,2}[-/]\d{4}"),  # DD-MM-YYYY or MM/DD/YYYY
    re.compile(r"^\d{4}\.\d{1,2}\.\d{1,2}"),  # YYYY.MM.DD
]


def _is_date_string(value: str) -> bool:
    """Check if a string matches common date patterns."""
    value = value.strip()
    for pattern in _DATE_PATTERNS:
        if pattern.match(value):
            return True
    return False


# --- Missing Value Detection ---


def detect_missing_values(columns: list[str], rows: list[dict[str, Any]]) -> MissingValueReport:
    """Detect missing values and their locations.

    Missing values are: None, empty string, NaN, or key not present.

    Args:
        columns: List of column names.
        rows: List of row dictionaries.

    Returns:
        MissingValueReport with positions and statistics.
    """
    positions: list[tuple[int, str]] = []
    per_column: dict[str, int] = {col: 0 for col in columns}
    total_missing = 0
    total_cells = len(rows) * len(columns) if columns else 0

    for row_idx, row in enumerate(rows):
        for col in columns:
            value = row.get(col)
            if _is_missing(value):
                positions.append((row_idx, col))
                per_column[col] += 1
                total_missing += 1

    overall_pct = (total_missing / total_cells * 100) if total_cells > 0 else 0.0

    return MissingValueReport(
        positions=positions,
        per_column=per_column,
        total_missing=total_missing,
        total_cells=total_cells,
        overall_percentage=round(overall_pct, 2),
    )


def _is_missing(value: Any) -> bool:
    """Check if a value is considered missing."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    if isinstance(value, float):
        import math
        if math.isnan(value):
            return True
    return False


# --- Format Error Detection ---


def detect_format_errors(content: bytes, file_format: str) -> list[FormatError]:
    """Detect format errors in the file content.

    Args:
        content: Raw file bytes.
        file_format: File format string (e.g., "csv", "xlsx", "json").

    Returns:
        List of FormatError instances with row/column info.
    """
    fmt = file_format.lstrip(".").lower()

    if fmt == "csv":
        return _detect_csv_errors(content)
    elif fmt in ("xlsx", "xls"):
        return _detect_excel_errors(content)
    elif fmt == "json":
        return _detect_json_errors(content)
    else:
        return [FormatError(row=0, column=None, description=f"Unsupported format: {file_format}")]


def _detect_csv_errors(content: bytes) -> list[FormatError]:
    """Detect CSV format errors: encoding issues, inconsistent column counts."""
    errors: list[FormatError] = []

    # Try decoding
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        errors.append(FormatError(row=0, column=None, description="File encoding is not valid UTF-8"))
        try:
            text = content.decode("latin-1")
        except Exception:
            return errors

    lines = text.splitlines()
    if not lines:
        return errors

    # Parse header to get expected column count
    try:
        reader = csv.reader(io.StringIO(lines[0]))
        header = next(reader)
        expected_cols = len(header)
    except Exception:
        errors.append(FormatError(row=1, column=None, description="Cannot parse CSV header"))
        return errors

    # Check each row for inconsistent column counts
    for line_idx, line in enumerate(lines[1:], start=2):
        if not line.strip():
            continue
        try:
            reader = csv.reader(io.StringIO(line))
            row_fields = next(reader)
            if len(row_fields) != expected_cols:
                errors.append(
                    FormatError(
                        row=line_idx,
                        column=None,
                        description=f"Inconsistent column count: expected {expected_cols}, got {len(row_fields)}",
                    )
                )
        except csv.Error as e:
            errors.append(FormatError(row=line_idx, column=None, description=f"CSV parse error: {str(e)}"))

    return errors


def _detect_excel_errors(content: bytes) -> list[FormatError]:
    """Detect Excel format errors: formula errors, merged cells."""
    errors: list[FormatError] = []

    try:
        import openpyxl
    except ImportError:
        errors.append(FormatError(row=0, column=None, description="openpyxl not available"))
        return errors

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        errors.append(FormatError(row=0, column=None, description=f"Cannot open Excel file: {str(e)}"))
        return errors

    ws = wb.active
    if ws is None:
        wb.close()
        return errors

    # Check for merged cells
    if ws.merged_cells.ranges:
        for merged_range in ws.merged_cells.ranges:
            errors.append(
                FormatError(
                    row=merged_range.min_row,
                    column=None,
                    description=f"Merged cells detected: {str(merged_range)}",
                )
            )

    # Check for error values in cells
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str):
                # Excel error values
                if cell.value in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#N/A", "#NUM!"):
                    col_letter = cell.column_letter if hasattr(cell, "column_letter") else str(cell.column)
                    errors.append(
                        FormatError(
                            row=row_idx,
                            column=col_letter,
                            description=f"Formula error: {cell.value}",
                        )
                    )

    wb.close()
    return errors


def _detect_json_errors(content: bytes) -> list[FormatError]:
    """Detect JSON format errors: malformed entries, inconsistent schemas."""
    errors: list[FormatError] = []

    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        errors.append(FormatError(row=0, column=None, description="File encoding is not valid UTF-8"))
        return errors

    try:
        data = json.loads(text)
    except json.JSONDecodeError as e:
        errors.append(FormatError(row=e.lineno or 0, column=None, description=f"JSON parse error: {e.msg}"))
        return errors

    # Get the records list
    records: list[Any] = []
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict) and "data" in data and isinstance(data["data"], list):
        records = data["data"]
    elif isinstance(data, dict):
        records = [data]

    if not records:
        return errors

    # Check for non-dict entries
    for idx, entry in enumerate(records):
        if not isinstance(entry, dict):
            errors.append(
                FormatError(
                    row=idx + 1,
                    column=None,
                    description=f"Entry is not an object (got {type(entry).__name__})",
                )
            )

    # Check for inconsistent schemas (compare keys to first record)
    dict_records = [r for r in records if isinstance(r, dict)]
    if dict_records:
        reference_keys = set(dict_records[0].keys())
        for idx, record in enumerate(dict_records[1:], start=2):
            current_keys = set(record.keys())
            missing_keys = reference_keys - current_keys
            extra_keys = current_keys - reference_keys
            if missing_keys:
                errors.append(
                    FormatError(
                        row=idx,
                        column=None,
                        description=f"Missing keys compared to first record: {sorted(missing_keys)}",
                    )
                )
            if extra_keys:
                errors.append(
                    FormatError(
                        row=idx,
                        column=None,
                        description=f"Extra keys compared to first record: {sorted(extra_keys)}",
                    )
                )

    return errors


# --- Quality Summary Generation ---


def generate_quality_summary(columns: list[str], rows: list[dict[str, Any]]) -> QualitySummary:
    """Generate a comprehensive data quality summary.

    Args:
        columns: List of column names.
        rows: List of row dictionaries.

    Returns:
        QualitySummary with overall and per-column quality metrics.
    """
    total_rows = len(rows)
    total_columns = len(columns)
    total_cells = total_rows * total_columns if total_columns > 0 else 0

    # Detect types and missing values
    col_types = detect_column_types(columns, rows)
    missing_report = detect_missing_values(columns, rows)

    # Build per-column quality info
    column_infos: list[ColumnQualityInfo] = []
    for col in columns:
        values = [row.get(col) for row in rows]
        non_missing_values = [v for v in values if not _is_missing(v)]
        missing_count = missing_report.per_column.get(col, 0)
        missing_pct = (missing_count / total_rows * 100) if total_rows > 0 else 0.0
        unique_count = len(set(str(v) for v in non_missing_values)) if non_missing_values else 0

        # Sample values (up to 5 unique non-missing values)
        seen = set()
        sample_values: list[Any] = []
        for v in non_missing_values:
            key = str(v)
            if key not in seen and len(sample_values) < 5:
                seen.add(key)
                sample_values.append(v)

        column_infos.append(
            ColumnQualityInfo(
                name=col,
                dtype=col_types.get(col, "text"),
                missing_count=missing_count,
                missing_percentage=round(missing_pct, 2),
                unique_count=unique_count,
                sample_values=sample_values,
            )
        )

    completeness = ((total_cells - missing_report.total_missing) / total_cells * 100) if total_cells > 0 else 100.0

    return QualitySummary(
        total_rows=total_rows,
        total_columns=total_columns,
        missing_value_percentage=missing_report.overall_percentage,
        completeness_score=round(completeness, 2),
        columns=column_infos,
    )
